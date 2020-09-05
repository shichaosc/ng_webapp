from django.core.management.base import BaseCommand
from course.models import CourseLesson, CourseUnit, CourseEdition, CourseInfo
import os
import openpyxl

'''配置高级版国际版小班课课程'''

class Command(BaseCommand):

    def add_arguments(self, parser):

        parser.add_argument('--file_name',
                            dest='file_name',
                            default='')

        parser.add_argument('--edition_name',
                            dest='edition_name',
                            default='')

        parser.add_argument('--sheet_name',
                            dest='sheet_name',
                            default='')
        parser.add_argument('--sheet_name2',
                            dest='sheet_name2',
                            default='')

    def handle(self, *args, **kwargs):
        file_name = kwargs.get('file_name')
        edition_name = kwargs.get('edition_name')
        sheet_name = kwargs.get('sheet_name', None)
        sheet_name2 = kwargs.get('sheet_name2', None)

        path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        path = os.path.join(path, 'files')
        file_path = os.path.join(path, file_name)
        print(file_path)
        # 读取excle
        wb = openpyxl.load_workbook(file_path)
        # 选择sheet
        if sheet_name:
            sh = wb[sheet_name]
            # ce = sh.cell(row=1, column=1)  # 读取第一行，
            course_edition = CourseEdition.objects.filter(edition_name=edition_name).first()
            if not course_edition:
                course_edition = CourseEdition()
                course_edition.edition_name = edition_name
                course_edition.description = ''
                course_edition.save()
            else:
                CourseInfo.objects.filter(course_edition_id=course_edition.id).all().delete()
            for cases in list(sh.rows)[1:]:
                course_name = str(cases[0].value).strip()
                course_level = str(cases[1].value).strip()
                lesson_no = cases[2].value
                unit_no = cases[3].value
                unit_lesson_no = cases[4].value
                report = cases[5].value

                course_info = CourseInfo.objects.filter(course_name=course_name, course_level=course_level, course_edition_id=course_edition.id).first()

                if not course_info:
                    course_info = CourseInfo()
                    course_info.course_name = course_name
                    course_info.course_edition_id = course_edition.id
                    course_info.course_level = course_level
                    course_info.save()

                course_lesson = CourseLesson()
                course_lesson.lesson_no = lesson_no
                course_lesson.course_id = course_info.id
                course_lesson.unit_lesson_no = unit_lesson_no
                course_lesson.unit_no = unit_no
                course_lesson.unit_report = report
                course_lesson.save()

                course_unit = CourseUnit.objects.filter(course_id=course_info.id, unit_no=unit_no).first()
                if not course_unit:
                    course_unit = CourseUnit()
                    course_unit.course_id = course_info.id
                    course_unit.unit_no = unit_no
                    course_unit.first_lesson_no = lesson_no
                    course_unit.last_lesson_no = lesson_no
                    course_unit.save()
                else:
                    course_unit.last_lesson_no = lesson_no
                    course_unit.save(update_fields=['last_lesson_no'])
        if sheet_name2:
            sh = wb[sheet_name]
            # ce = sh.cell(row=1, column=1)  # 读取第一行，
            course_edition = CourseEdition.objects.filter(edition_name=edition_name).first()
            if not course_edition:
                course_edition = CourseEdition()
                course_edition.edition_name = edition_name
                course_edition.description = ''
                course_edition.save()
            else:
                CourseInfo.objects.filter(course_edition_id=course_edition.id).all().delete()
            for cases in list(sh.rows)[1:]:
                course_name = str(cases[0].value).strip()
                course_level = str(cases[1].value).strip()
                lesson_no = cases[2].value
                unit_no = cases[3].value
                unit_lesson_no = cases[4].value
                report = cases[5].value

                course_info = CourseInfo.objects.filter(course_name=course_name, course_level=course_level, course_edition_id=course_edition.id).first()

                if not course_info:
                    course_info = CourseInfo()
                    course_info.course_name = course_name
                    course_info.course_edition_id = course_edition.id
                    course_info.course_level = course_level
                    course_info.save()

                course_lesson = CourseLesson()
                course_lesson.lesson_no = lesson_no
                course_lesson.course_id = course_info.id
                course_lesson.unit_lesson_no = unit_lesson_no
                course_lesson.unit_no = unit_no
                course_lesson.unit_report = report
                course_lesson.save()

                course_unit = CourseUnit.objects.filter(course_id=course_info.id, unit_no=unit_no).first()
                if not course_unit:
                    course_unit = CourseUnit()
                    course_unit.course_id = course_info.id
                    course_unit.unit_no = unit_no
                    course_unit.first_lesson_no = lesson_no
                    course_unit.last_lesson_no = lesson_no
                    course_unit.save()
                else:
                    course_unit.last_lesson_no = lesson_no
                    course_unit.save(update_fields=['last_lesson_no'])
        wb.close()
